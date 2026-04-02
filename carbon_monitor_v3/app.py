from flask import Flask, render_template, request, redirect, url_for, jsonify, send_from_directory
import sqlite3
import os
import csv
import io
from datetime import datetime
from werkzeug.utils import secure_filename

try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

app = Flask(__name__)
DATABASE    = 'carbon.db'
UPLOAD_FOLDER   = 'uploads'
ALLOWED_EXTENSIONS = {'csv', 'xlsx'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ── Emission Factors (kg CO2 per unit) ──────────────────────────
EMISSION_FACTORS = {
    'electricity': 0.82,   # kg CO2 per kWh
    'fuel':        2.31,   # kg CO2 per litre
    'transport':   0.21,   # kg CO2 per km
    'waste':       0.50,   # kg CO2 per kg
}

# ── Database Setup ───────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS emissions (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            date        TEXT    NOT NULL,
            activity    TEXT    NOT NULL,
            value       REAL    NOT NULL,
            unit        TEXT    NOT NULL,
            co2_kg      REAL    NOT NULL,
            notes       TEXT
        )
    ''')
    conn.commit()
    conn.close()

# ── Routes ───────────────────────────────────────────────────────

@app.route('/')
def index():
    return redirect(url_for('dashboard'))


@app.route('/dashboard')
def dashboard():
    conn = get_db()

    # Total CO2
    total = conn.execute('SELECT SUM(co2_kg) as total FROM emissions').fetchone()['total'] or 0

    # CO2 by activity
    by_activity = conn.execute('''
        SELECT activity, SUM(co2_kg) as total
        FROM emissions GROUP BY activity
    ''').fetchall()

    # Monthly trend (last 6 months)
    monthly = conn.execute('''
        SELECT strftime('%Y-%m', date) as month, SUM(co2_kg) as total
        FROM emissions
        GROUP BY month
        ORDER BY month DESC
        LIMIT 6
    ''').fetchall()

    # Recent 5 records
    recent = conn.execute('''
        SELECT * FROM emissions ORDER BY date DESC LIMIT 5
    ''').fetchall()

    conn.close()

    return render_template('dashboard.html',
        total=round(total, 2),
        by_activity=by_activity,
        monthly=list(reversed(monthly)),
        recent=recent
    )


@app.route('/add', methods=['GET', 'POST'])
def add_emission():
    if request.method == 'POST':
        activity = request.form['activity']
        value    = float(request.form['value'])
        date     = request.form['date']
        notes    = request.form.get('notes', '')

        factor   = EMISSION_FACTORS.get(activity, 0)
        co2_kg   = round(value * factor, 4)

        units = {
            'electricity': 'kWh',
            'fuel':        'Litres',
            'transport':   'km',
            'waste':       'kg',
        }

        conn = get_db()
        conn.execute('''
            INSERT INTO emissions (date, activity, value, unit, co2_kg, notes)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (date, activity, value, units[activity], co2_kg, notes))
        conn.commit()
        conn.close()

        return redirect(url_for('dashboard'))

    today = datetime.today().strftime('%Y-%m-%d')
    return render_template('add.html', today=today, factors=EMISSION_FACTORS)


@app.route('/records')
def records():
    conn = get_db()
    all_records = conn.execute(
        'SELECT * FROM emissions ORDER BY date DESC'
    ).fetchall()
    conn.close()
    return render_template('records.html', records=all_records)


@app.route('/delete/<int:record_id>', methods=['POST'])
def delete_record(record_id):
    conn = get_db()
    conn.execute('DELETE FROM emissions WHERE id = ?', (record_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('records'))


@app.route('/api/chart-data')
def chart_data():
    conn = get_db()

    by_activity = conn.execute('''
        SELECT activity, SUM(co2_kg) as total
        FROM emissions GROUP BY activity
    ''').fetchall()

    monthly = conn.execute('''
        SELECT strftime('%Y-%m', date) as month, SUM(co2_kg) as total
        FROM emissions GROUP BY month ORDER BY month ASC LIMIT 6
    ''').fetchall()

    conn.close()

    return jsonify({
        'by_activity': {
            'labels': [r['activity'].capitalize() for r in by_activity],
            'values': [round(r['total'], 2) for r in by_activity],
        },
        'monthly': {
            'labels': [r['month'] for r in monthly],
            'values': [round(r['total'], 2) for r in monthly],
        }
    })


# ── Helpers ─────────────────────────────────────────────────────
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def process_rows(rows):
    """Validate and insert a list of dicts with keys: date, activity, value, notes"""
    units = {'electricity': 'kWh', 'fuel': 'Litres', 'transport': 'km', 'waste': 'kg'}
    inserted, skipped, errors = 0, 0, []

    conn = get_db()
    for i, row in enumerate(rows, start=2):
        try:
            date     = str(row.get('date', '')).strip()
            activity = str(row.get('activity', '')).strip().lower()
            value    = float(str(row.get('value', '')).strip())
            notes    = str(row.get('notes', '') or '').strip()

            if not date or activity not in EMISSION_FACTORS or value <= 0:
                skipped += 1
                errors.append(f"Row {i}: invalid data — date={date}, activity={activity}, value={value}")
                continue

            co2_kg = round(value * EMISSION_FACTORS[activity], 4)
            conn.execute('''
                INSERT INTO emissions (date, activity, value, unit, co2_kg, notes)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (date, activity, value, units[activity], co2_kg, notes))
            inserted += 1
        except Exception as e:
            skipped += 1
            errors.append(f"Row {i}: {str(e)}")

    conn.commit()
    conn.close()
    return inserted, skipped, errors


@app.route('/upload', methods=['GET', 'POST'])
def upload_file():
    result = None

    if request.method == 'POST':
        if 'file' not in request.files:
            result = {'status': 'error', 'message': 'No file selected.'}
        else:
            file = request.files['file']
            if file.filename == '':
                result = {'status': 'error', 'message': 'No file chosen.'}
            elif not allowed_file(file.filename):
                result = {'status': 'error', 'message': 'Only .csv and .xlsx files are allowed.'}
            else:
                filename  = secure_filename(file.filename)
                ext       = filename.rsplit('.', 1)[1].lower()
                filepath  = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                rows = []

                try:
                    if ext == 'csv':
                        with open(filepath, newline='', encoding='utf-8-sig') as f:
                            reader = csv.DictReader(f)
                            # Normalise column names to lowercase stripped
                            for r in reader:
                                rows.append({k.strip().lower(): v for k, v in r.items()})

                    elif ext == 'xlsx':
                        if not EXCEL_SUPPORT:
                            result = {'status': 'error', 'message': 'openpyxl not installed. Run: pip install openpyxl'}
                        else:
                            wb = openpyxl.load_workbook(filepath)
                            ws = wb.active
                            headers = [str(c.value).strip().lower() if c.value else '' for c in ws[1]]
                            for row in ws.iter_rows(min_row=2, values_only=True):
                                if any(v is not None for v in row):
                                    rows.append(dict(zip(headers, row)))

                    if result is None:
                        inserted, skipped, errors = process_rows(rows)
                        result = {
                            'status':   'success' if inserted > 0 else 'warning',
                            'inserted': inserted,
                            'skipped':  skipped,
                            'errors':   errors[:5],   # show max 5 errors
                            'filename': filename,
                        }
                except Exception as e:
                    result = {'status': 'error', 'message': f'Could not read file: {str(e)}'}

                finally:
                    if os.path.exists(filepath):
                        os.remove(filepath)   # clean up

    return render_template('upload.html', result=result)


@app.route('/recommendations')
def recommendations():
    conn = get_db()

    # Total per activity
    by_activity = conn.execute('''
        SELECT activity, SUM(co2_kg) as total, COUNT(*) as count, AVG(co2_kg) as avg
        FROM emissions GROUP BY activity
    ''').fetchall()

    # Monthly totals (last 6)
    monthly = conn.execute('''
        SELECT strftime('%Y-%m', date) as month, SUM(co2_kg) as total
        FROM emissions GROUP BY month ORDER BY month DESC LIMIT 6
    ''').fetchall()

    # Grand total
    grand_total = conn.execute('SELECT SUM(co2_kg) as t FROM emissions').fetchone()['t'] or 0

    # Highest single record
    highest = conn.execute('''
        SELECT * FROM emissions ORDER BY co2_kg DESC LIMIT 1
    ''').fetchone()

    conn.close()

    # ── Build recommendations ────────────────────────────────────
    recs = []
    totals = {r['activity']: r['total'] for r in by_activity}
    counts = {r['activity']: r['count']  for r in by_activity}

    # 1. Electricity
    elec = totals.get('electricity', 0)
    if elec > 0:
        if elec > 200:
            recs.append({
                'category': 'electricity',
                'priority': 'high',
                'icon': '⚡',
                'title': 'High Electricity Consumption Detected',
                'detail': f'Electricity accounts for {round(elec, 1)} kg CO₂ — your largest emission source.',
                'actions': [
                    'Switch to LED lighting across all laboratory areas',
                    'Install smart power strips to eliminate standby power usage',
                    'Schedule equipment shutdowns during non-working hours',
                    'Consider renewable energy sources (solar panels) for the facility',
                ]
            })
        elif elec > 80:
            recs.append({
                'category': 'electricity',
                'priority': 'medium',
                'icon': '⚡',
                'title': 'Moderate Electricity Usage — Room for Improvement',
                'detail': f'Electricity emissions: {round(elec, 1)} kg CO₂. Optimisation can reduce this significantly.',
                'actions': [
                    'Enable energy-saving mode on all computers and monitors',
                    'Use natural lighting where possible during daytime',
                    'Audit and replace old energy-intensive lab equipment',
                ]
            })
        else:
            recs.append({
                'category': 'electricity',
                'priority': 'low',
                'icon': '⚡',
                'title': 'Electricity Usage is Under Control',
                'detail': f'Electricity emissions: {round(elec, 1)} kg CO₂. Keep maintaining current practices.',
                'actions': [
                    'Continue monitoring monthly consumption',
                    'Maintain equipment servicing schedules for efficiency',
                ]
            })

    # 2. Transport
    trans = totals.get('transport', 0)
    if trans > 0:
        if trans > 100:
            recs.append({
                'category': 'transport',
                'priority': 'high',
                'icon': '🚗',
                'title': 'Transportation Emissions are Significant',
                'detail': f'Transport contributes {round(trans, 1)} kg CO₂ across {counts.get("transport", 0)} trips.',
                'actions': [
                    'Introduce carpooling or shared transport for staff',
                    'Shift short meetings to video conferencing instead of travel',
                    'Plan and consolidate field visits to reduce trip frequency',
                    'Consider switching to electric or hybrid vehicles for facility use',
                ]
            })
        else:
            recs.append({
                'category': 'transport',
                'priority': 'medium',
                'icon': '🚗',
                'title': 'Transportation Emissions Can Be Reduced',
                'detail': f'Transport emissions: {round(trans, 1)} kg CO₂.',
                'actions': [
                    'Use video calls to replace non-essential travel',
                    'Encourage public transport use with incentives',
                ]
            })

    # 3. Fuel
    fuel = totals.get('fuel', 0)
    if fuel > 0:
        if fuel > 100:
            recs.append({
                'category': 'fuel',
                'priority': 'high',
                'icon': '⛽',
                'title': 'Fuel Consumption Needs Attention',
                'detail': f'Fuel usage contributes {round(fuel, 1)} kg CO₂ to total emissions.',
                'actions': [
                    'Audit generator usage — run only when essential',
                    'Maintain equipment regularly to improve fuel efficiency',
                    'Explore solar or battery backup to reduce generator dependency',
                    'Track and log fuel usage per equipment for accountability',
                ]
            })
        else:
            recs.append({
                'category': 'fuel',
                'priority': 'low',
                'icon': '⛽',
                'title': 'Fuel Usage is Relatively Low',
                'detail': f'Fuel emissions: {round(fuel, 1)} kg CO₂. Continue good practices.',
                'actions': [
                    'Keep fuel logs up to date for audit purposes',
                    'Service engines regularly to maintain efficiency',
                ]
            })

    # 4. Waste
    waste = totals.get('waste', 0)
    if waste > 0:
        if waste > 50:
            recs.append({
                'category': 'waste',
                'priority': 'high',
                'icon': '🗑️',
                'title': 'Waste Generation is Above Recommended Levels',
                'detail': f'Waste contributes {round(waste, 1)} kg CO₂. Reduction is highly recommended.',
                'actions': [
                    'Implement a waste segregation programme (recyclables vs general)',
                    'Reduce single-use plastics in laboratory and office areas',
                    'Set up composting for organic waste',
                    'Partner with certified waste recycling agencies',
                ]
            })
        else:
            recs.append({
                'category': 'waste',
                'priority': 'low',
                'icon': '🗑️',
                'title': 'Waste Levels are Manageable',
                'detail': f'Waste emissions: {round(waste, 1)} kg CO₂.',
                'actions': [
                    'Continue waste segregation practices',
                    'Educate staff on reduce-reuse-recycle principles',
                ]
            })

    # 5. Trend-based recommendation
    if len(monthly) >= 2:
        latest  = monthly[0]['total']
        prev    = monthly[1]['total']
        change  = round(((latest - prev) / prev) * 100, 1) if prev > 0 else 0
        if change > 15:
            recs.append({
                'category': 'trend',
                'priority': 'high',
                'icon': '📈',
                'title': f'Emissions Rose {change}% This Month — Investigate',
                'detail': f'Last month: {round(prev,1)} kg → This month: {round(latest,1)} kg CO₂.',
                'actions': [
                    'Identify which activity spiked this month',
                    'Review recent changes in equipment usage or operations',
                    'Set a monthly reduction target of at least 10%',
                ]
            })
        elif change < -10:
            recs.append({
                'category': 'trend',
                'priority': 'low',
                'icon': '📉',
                'title': f'Emissions Dropped {abs(change)}% — Great Progress!',
                'detail': f'Reduction from {round(prev,1)} kg to {round(latest,1)} kg CO₂ this month.',
                'actions': [
                    'Document what changes led to this improvement',
                    'Share best practices with other departments',
                    'Set a new target to sustain this reduction',
                ]
            })

    # 6. Overall health score (simple)
    if grand_total > 0:
        if grand_total < 200:
            health = {'score': 90, 'label': 'Excellent', 'color': '2c5f2d'}
        elif grand_total < 500:
            health = {'score': 70, 'label': 'Good',      'color': '1a5276'}
        elif grand_total < 1000:
            health = {'score': 45, 'label': 'Moderate',  'color': 'b85c00'}
        else:
            health = {'score': 20, 'label': 'Critical',  'color': 'c0392b'}
    else:
        health = None

    # Sort: high → medium → low
    priority_order = {'high': 0, 'medium': 1, 'low': 2}
    recs.sort(key=lambda x: priority_order.get(x['priority'], 3))

    return render_template('recommendations.html',
        recs=recs,
        totals=totals,
        grand_total=round(grand_total, 2),
        health=health,
        monthly=list(reversed(monthly)),
        highest=highest,
    )



def download_template(fmt):
    if fmt in ('csv', 'xlsx'):
        return send_from_directory('sample', f'sample_template.{fmt}', as_attachment=True)
    return "Invalid format", 404


# ── Run ──────────────────────────────────────────────────────────
if __name__ == '__main__':
    init_db()
    print("\n✅  Carbon Monitor running at: http://127.0.0.1:5000\n")
    app.run(debug=True)
